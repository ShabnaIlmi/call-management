import streamlit as st
import mysql.connector
import bcrypt
import pandas as pd
from datetime import datetime, timedelta
import time
import io
import base64
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import re  # For email extraction

# Database setup modifications
def init_db():
    """Initialize database with customer table if not exists"""
    try:
        db = connect_db()
        cursor = db.cursor()
        
        # Create customers table if not exists
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            email VARCHAR(255) UNIQUE,
            phone VARCHAR(50),
            first_contact DATETIME,
            last_contact DATETIME,
            notes TEXT
        )
        """)
        
        # Add email column to messages if not exists
        cursor.execute("SHOW COLUMNS FROM messages LIKE 'customer_email'")
        if not cursor.fetchone():
            cursor.execute("ALTER TABLE messages ADD COLUMN customer_email VARCHAR(255)")
        
        db.commit()
        cursor.close()
        db.close()
    except Exception as e:
        st.error(f"Database initialization error: {e}")

# Email extraction function
def extract_email(text):
    """Extract email from message text using regex"""
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    match = re.search(email_pattern, text)
    return match.group(0) if match else None

# Customer management functions
def get_or_create_customer(sender_name, message_text):
    """Get existing customer or create new one"""
    try:
        db = connect_db()
        cursor = db.cursor(dictionary=True)
        
        # Extract email from message
        email = extract_email(message_text)
        
        if not email:
            return None
            
        # Check if customer exists
        cursor.execute("SELECT * FROM customers WHERE email = %s", (email,))
        customer = cursor.fetchone()
        
        if not customer:
            # Create new customer
            cursor.execute("""
            INSERT INTO customers 
            (name, email, first_contact, last_contact) 
            VALUES (%s, %s, %s, %s)
            """, (sender_name, email, datetime.now(), datetime.now()))
            db.commit()
            customer_id = cursor.lastrowid
        else:
            # Update last contact time
            cursor.execute("""
            UPDATE customers SET last_contact = %s 
            WHERE email = %s
            """, (datetime.now(), email))
            db.commit()
            customer_id = customer['id']
        
        cursor.close()
        db.close()
        return customer_id
    except Exception as e:
        st.error(f"Customer management error: {e}")
        return None

# Modified message processing
def process_incoming_message(sender, message):
    """Process new incoming message with customer tracking"""
    try:
        db = connect_db()
        cursor = db.cursor()
        
        # Handle customer record
        customer_id = get_or_create_customer(sender, message)
        email = extract_email(message)
        
        # Store message with customer reference
        cursor.execute("""
        INSERT INTO messages 
        (sender, message, timestamp, customer_email) 
        VALUES (%s, %s, %s, %s)
        """, (sender, message, datetime.now(), email))
        
        db.commit()
        cursor.close()
        db.close()
        return True
    except Exception as e:
        st.error(f"Message processing error: {e}")
        return False

# Enhanced email function
def send_completion_notifications(task_details, completion_note, assigned_to):
    """Send emails to both admin and customer"""
    try:
        # Email to admin (info@busy.lk)
        admin_msg = MIMEMultipart()
        admin_msg['From'] = EMAIL_ADDRESS
        admin_msg['To'] = RECIPIENT_EMAIL
        admin_msg['Subject'] = f"Task Completed: {task_details.get('sender', 'Unknown')}"
        
        admin_body = f"""
        <h2>Task Completion Notification</h2>
        <p>A task has been marked as completed by {assigned_to}.</p>
        <h3>Customer Details:</h3>
        <ul>
            <li><strong>Name:</strong> {task_details.get('sender', 'Unknown')}</li>
            <li><strong>Email:</strong> {task_details.get('customer_email', 'Not provided')}</li>
        </ul>
        <h3>Task Details:</h3>
        <p>{task_details.get('message', 'No message')}</p>
        <h3>Completion Note:</h3>
        <p>{completion_note}</p>
        """
        admin_msg.attach(MIMEText(admin_body, 'html'))
        
        # Email to customer
        customer_email = task_details.get('customer_email')
        if customer_email:
            customer_msg = MIMEMultipart()
            customer_msg['From'] = EMAIL_ADDRESS
            customer_msg['To'] = customer_email
            customer_msg['Subject'] = f"Your request has been completed - {task_details.get('sender', 'Unknown')}"
            
            customer_body = f"""
            <h2>Request Completion Notification</h2>
            <p>Dear {task_details.get('sender', 'Customer')},</p>
            <p>Your request has been successfully completed by our team.</p>
            <h3>Completion Details:</h3>
            <p><strong>Request:</strong> {task_details.get('message', 'No message')}</p>
            <p><strong>Completion Note:</strong> {completion_note}</p>
            <p><strong>Completed by:</strong> {assigned_to}</p>
            <p>Thank you for your business!</p>
            """
            customer_msg.attach(MIMEText(customer_body, 'html'))
        
        # Send emails
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(admin_msg)
            if customer_email:
                server.send_message(customer_msg)
        
        return True
    except Exception as e:
        st.error(f"Email sending error: {e}")
        return False

# Add these configuration variables at the top of your script
SMTP_SERVER = "smtp.gmail.com"  
SMTP_PORT = 587
EMAIL_ADDRESS = "info@busy.lk"  
EMAIL_PASSWORD = "your_app_password"  
RECIPIENT_EMAIL = "info@busy.lk"  

def send_completion_email(task_details, completion_note, assigned_to):
    """Send an email notification when a task is completed"""
    try:
        # Create the email message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = RECIPIENT_EMAIL
        msg['Subject'] = f"Task Completed: {task_details.get('sender', 'Unknown')} - {assigned_to}"
        
        # Create the email body
        body = f"""
        <h2>Task Completion Notification</h2>
        <p>A task has been marked as completed in the management system.</p>
        
        <h3>Task Details:</h3>
        <ul>
            <li><strong>From:</strong> {task_details.get('sender', 'Unknown')}</li>
            <li><strong>Original Message:</strong> {task_details.get('message', 'No message')}</li>
            <li><strong>Received at:</strong> {task_details.get('timestamp', 'Unknown time')}</li>
        </ul>
        
        <h3>Completion Details:</h3>
        <ul>
            <li><strong>Completed by:</strong> {assigned_to}</li>
            <li><strong>Completion Time:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</li>
            <li><strong>Time Taken:</strong> {task_details.get('time_taken', 'N/A')} minutes</li>
        </ul>
        
        <h3>Completion Note:</h3>
        <p>{completion_note}</p>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        # Connect to the SMTP server and send the email
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(msg)
        
        return True
    except Exception as e:
        st.error(f"Error sending email: {e}")
        return False

def connect_db():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="management_system"
    )

def authenticate(username, password):
    try:
        db = connect_db()
        cursor = db.cursor()
        # Using the exact column name 'password_hash' from your table structure
        cursor.execute("SELECT password_hash FROM users WHERE username = %s", (username,))
        result = cursor.fetchone()
        cursor.close()
        db.close()
        
        if result and bcrypt.checkpw(password.encode('utf-8'), result[0].encode('utf-8')):
            return True
        return False
    except Exception as e:
        st.error(f"DB Error: {e}")
        return False

def update_message_status(message_id, status, username, completion_note=None):
    try:
        db = connect_db()
        cursor = db.cursor()
        
        # First, let's check if assigned_to column exists
        try:
            cursor.execute("SHOW COLUMNS FROM messages LIKE 'assigned_to'")
            assigned_to_exists = cursor.fetchone() is not None
            
            if not assigned_to_exists:
                # Add assigned_to column if it doesn't exist
                cursor.execute("ALTER TABLE messages ADD COLUMN assigned_to VARCHAR(100)")
        except:
            # If there's an error, we'll add the column anyway and catch any error
            try:
                cursor.execute("ALTER TABLE messages ADD COLUMN assigned_to VARCHAR(100)")
            except:
                pass
        
        # Same check for updated_at column
        try:
            cursor.execute("SHOW COLUMNS FROM messages LIKE 'updated_at'")
            updated_at_exists = cursor.fetchone() is not None
            
            if not updated_at_exists:
                # Add updated_at column if it doesn't exist
                cursor.execute("ALTER TABLE messages ADD COLUMN updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        except:
            try:
                cursor.execute("ALTER TABLE messages ADD COLUMN updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
            except:
                pass
        
        # Check for completion_note column
        try:
            cursor.execute("SHOW COLUMNS FROM messages LIKE 'completion_note'")
            completion_note_exists = cursor.fetchone() is not None
            
            if not completion_note_exists:
                # Add completion_note column if it doesn't exist
                cursor.execute("ALTER TABLE messages ADD COLUMN completion_note TEXT")
        except:
            try:
                cursor.execute("ALTER TABLE messages ADD COLUMN completion_note TEXT")
            except:
                pass
        
        # Check for time_taken column
        try:
            cursor.execute("SHOW COLUMNS FROM messages LIKE 'time_taken'")
            time_taken_exists = cursor.fetchone() is not None
            
            if not time_taken_exists:
                # Add time_taken column if it doesn't exist
                cursor.execute("ALTER TABLE messages ADD COLUMN time_taken INT")
        except:
            try:
                cursor.execute("ALTER TABLE messages ADD COLUMN time_taken INT")
            except:
                pass
        
        # Calculate time taken if status is changing to completed
        time_taken_mins = None
        if status == "completed":
            # Get the timestamp when the task was set to ongoing
            cursor.execute("SELECT updated_at FROM messages WHERE id = %s AND status = 'ongoing'", (message_id,))
            start_time_result = cursor.fetchone()
            
            if start_time_result and start_time_result[0]:
                start_time = start_time_result[0]
                end_time = datetime.now()
                time_diff = end_time - start_time
                time_taken_mins = int(time_diff.total_seconds() / 60)
        
        # Now update the message status
        if completion_note is not None and time_taken_mins is not None:
            cursor.execute(
                "UPDATE messages SET status = %s, assigned_to = %s, updated_at = %s, completion_note = %s, time_taken = %s WHERE id = %s",
                (status, username, datetime.now(), completion_note, time_taken_mins, message_id)
            )
        elif completion_note is not None:
            cursor.execute(
                "UPDATE messages SET status = %s, assigned_to = %s, updated_at = %s, completion_note = %s WHERE id = %s",
                (status, username, datetime.now(), completion_note, message_id)
            )
        else:
            cursor.execute(
                "UPDATE messages SET status = %s, assigned_to = %s, updated_at = %s WHERE id = %s",
                (status, username, datetime.now(), message_id)
            )
        db.commit()
        cursor.close()
        db.close()
        return True
    except Exception as e:
        st.error(f"Error updating message status: {e}")
        return False

def get_message_count():
    try:
        db = connect_db()
        cursor = db.cursor()
        cursor.execute("SELECT COUNT(*) FROM messages WHERE status != 'ongoing' OR status IS NULL")
        count = cursor.fetchone()[0]
        cursor.close()
        db.close()
        return count
    except Exception as e:
        st.error(f"Error counting messages: {e}")
        return 0

def get_activity_data(date_range='today'):
    try:
        db = connect_db()
        cursor = db.cursor(dictionary=True)
        
        if date_range == 'today':
            today = datetime.now().date()
            query = """
                SELECT m.id, m.sender, m.message, m.status, m.assigned_to, 
                       m.timestamp, m.updated_at, m.completion_note, m.time_taken
                FROM messages m
                WHERE DATE(m.updated_at) = %s AND m.status = 'completed'
                ORDER BY m.updated_at DESC
            """
            cursor.execute(query, (today,))
        elif date_range == 'week':
            today = datetime.now().date()
            week_ago = today - timedelta(days=7)
            query = """
                SELECT m.id, m.sender, m.message, m.status, m.assigned_to, 
                       m.timestamp, m.updated_at, m.completion_note, m.time_taken
                FROM messages m
                WHERE DATE(m.updated_at) BETWEEN %s AND %s AND m.status = 'completed'
                ORDER BY m.updated_at DESC
            """
            cursor.execute(query, (week_ago, today))
        elif date_range == 'month':
            today = datetime.now().date()
            month_ago = today - timedelta(days=30)
            query = """
                SELECT m.id, m.sender, m.message, m.status, m.assigned_to, 
                       m.timestamp, m.updated_at, m.completion_note, m.time_taken
                FROM messages m
                WHERE DATE(m.updated_at) BETWEEN %s AND %s AND m.status = 'completed'
                ORDER BY m.updated_at DESC
            """
            cursor.execute(query, (month_ago, today))
        else:
            # Parse custom date range
            start_date, end_date = date_range.split('to')
            query = """
                SELECT m.id, m.sender, m.message, m.status, m.assigned_to, 
                       m.timestamp, m.updated_at, m.completion_note, m.time_taken
                FROM messages m
                WHERE DATE(m.updated_at) BETWEEN %s AND %s AND m.status = 'completed'
                ORDER BY m.updated_at DESC
            """
            cursor.execute(query, (start_date.strip(), end_date.strip()))
            
        activities = cursor.fetchall()
        cursor.close()
        db.close()
        return activities
    except Exception as e:
        st.error(f"Error fetching activity data: {e}")
        return []

def create_report_preview(activities):
    """Create a report preview for display"""
    if not activities:
        return pd.DataFrame()
    
    # Create DataFrame for display
    df = pd.DataFrame(activities)
    
    # Format the DataFrame for display
    df_display = df.copy()
    
    # Format date columns
    if 'timestamp' in df_display.columns:
        df_display['timestamp'] = pd.to_datetime(df_display['timestamp']).dt.strftime('%Y-%m-%d %H:%M:%S')
    if 'updated_at' in df_display.columns:
        df_display['updated_at'] = pd.to_datetime(df_display['updated_at']).dt.strftime('%Y-%m-%d %H:%M:%S')
    
    # Truncate long messages for display
    if 'message' in df_display.columns:
        df_display['message'] = df_display['message'].apply(lambda x: (x[:50] + '...') if isinstance(x, str) and len(x) > 50 else x)
    
    # Select columns for display
    display_cols = ['sender', 'message', 'assigned_to', 'updated_at', 'time_taken']
    display_cols = [col for col in display_cols if col in df_display.columns]
    
    # Rename columns for better readability
    column_names = {
        'sender': 'Sender',
        'message': 'Message',
        'assigned_to': 'Completed By',
        'updated_at': 'Completion Time',
        'time_taken': 'Minutes Taken'
    }
    
    df_display = df_display[display_cols].rename(columns={k: v for k, v in column_names.items() if k in display_cols})
    
    return df_display

def get_download_link(df, filename, filetype='xlsx', report_title=None):
    """Generate a download link for dataframe with enhanced formatting
    
    Args:
        df: Pandas DataFrame containing the report data
        filename: Name for the downloaded file
        filetype: File format ('xlsx' or 'pdf')
        report_title: Optional title for the report (defaults to 'Activity Report')
    
    Returns:
        HTML string with download link
    """
    if not report_title:
        report_title = f"Activity Report - {datetime.now().strftime('%Y-%m-%d')}"
        
    if filetype == 'xlsx':
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the dataframe to the Excel file
            df.to_excel(writer, index=False, sheet_name='Activity Report')
            
            # Get the openpyxl workbook and worksheet objects for formatting
            workbook = writer.book
            worksheet = writer.sheets['Activity Report']
            
            # Add basic formatting for header row
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Create styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
            centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Apply header formatting
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered_alignment
                cell.border = thin_border
            
            # Set column widths based on content
            for i, col in enumerate(df.columns):
                # Calculate width based on content
                max_len = max(
                    df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                    len(str(col))
                ) + 2  # Add a little extra space
                
                # Cap width at 50 characters
                col_width = min(max_len, 50)
                
                # Set column width
                column_letter = get_column_letter(i+1)
                worksheet.column_dimensions[column_letter].width = col_width
            
            # Insert a title row at the top
            worksheet.insert_rows(1)
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = report_title
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Merge cells for the title
            end_col = min(len(df.columns), 5)
            if end_col > 1:
                worksheet.merge_cells(f'A1:{get_column_letter(end_col)}1')
            
            # Adjust row height for title
            worksheet.row_dimensions[1].height = 25
            
        binary_data = output.getvalue()
        mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
    elif filetype == 'pdf':
        # Create a PDF bytes buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, 
                               topMargin=30, bottomMargin=30)
        elements = []
        
        # Add title with better styling
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.alignment = 1  # Center alignment
        elements.append(Paragraph(report_title, title_style))
        elements.append(Paragraph(" ", styles['Normal']))  # Add some space
        
        # Add timestamp
        timestamp_style = styles['Italic']
        timestamp_style.alignment = 1
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", timestamp_style))
        elements.append(Paragraph(" ", styles['Normal']))  # Add some space
        
        # Create table data from dataframe
        data = [df.columns.tolist()] + df.values.tolist()
        
        # Format timestamps and handle other data types
        for i in range(1, len(data)):
            for j in range(len(data[i])):
                # Handle timestamps
                if isinstance(data[i][j], pd.Timestamp) or isinstance(data[i][j], datetime):
                    data[i][j] = data[i][j].strftime('%Y-%m-%d %H:%M:%S')
                # Handle None/NaN values
                elif pd.isna(data[i][j]):
                    data[i][j] = "N/A"
                # Handle long strings - truncate for better PDF layout
                elif isinstance(data[i][j], str) and len(data[i][j]) > 70:
                    data[i][j] = data[i][j][:67] + "..."
        
        # Create table with better styling
        # Calculate column widths based on content
        col_widths = []
        for i in range(len(df.columns)):
            # Default width
            col_width = 80
            
            # Adjust column width based on header length
            header_width = len(str(df.columns[i])) * 5
            col_width = max(col_width, header_width)
            
            if i == 1 and 'message' in df.columns:
                # Message column gets more space
                col_width = 200
            
            col_widths.append(col_width)
            
        table = Table(data, colWidths=col_widths)
        
        # Enhanced table styling
        table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling - alternate row colors
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('BACKGROUND', (0, 2), (-1, 2), colors.lightgrey),
            ('BACKGROUND', (0, 4), (-1, 4), colors.lightgrey),
            ('BACKGROUND', (0, 6), (-1, 6), colors.lightgrey),
            ('BACKGROUND', (0, 8), (-1, 8), colors.lightgrey),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            
            # Grid lines
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Specific column alignments
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),  # First column left aligned
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Table borders
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
        ]))
        
        elements.append(table)
        
        # Add footer with page numbers
        def add_page_number(canvas, doc):
            page_num = canvas.getPageNumber()
            text = f"Page {page_num}"
            canvas.setFont("Helvetica", 9)
            canvas.drawRightString(letter[0]-30, 30, text)
        
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        
        binary_data = buffer.getvalue()
        mime_type = 'application/pdf'
    else:
        return None
    
    # Create a styled button as the download link
    b64 = base64.b64encode(binary_data).decode()
    icon = "📊" if filetype == "xlsx" else "📄"
    button_style = """
        display: inline-block;
        padding: 0.5em 1em;
        margin: 0.5em 0;
        border-radius: 5px;
        background-color: #4CAF50;
        color: white;
        text-decoration: none;
        font-weight: bold;
        text-align: center;
        cursor: pointer;
    """
    href = f'''
    <a href="data:{mime_type};base64,{b64}" download="{filename}" style="{button_style}">
        {icon} Download {filetype.upper()} Report
    </a>
    '''
    return href

def login():
    st.title("🔐 Login")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    
    if st.button("Login"):
        if authenticate(username, password):
            st.session_state.logged_in = True
            st.session_state.username = username
            # Initialize last_count to trigger first refresh
            st.session_state.last_count = -1
            st.success(f"Welcome, {username}!")
            st.rerun()
        else:
            st.error("Invalid login")

def dashboard():
    # Initialize session state variables
    if 'current_view' not in st.session_state:
        st.session_state.current_view = "tasks"  # Default view
    
    if 'last_refresh' not in st.session_state:
        st.session_state.last_refresh = time.time()
        
    if 'last_count' not in st.session_state:
        st.session_state.last_count = get_message_count()
    
    if 'show_completion_form' not in st.session_state:
        st.session_state.show_completion_form = False
    
    if 'completion_task' not in st.session_state:
        st.session_state.completion_task = None
    
    # Set up layout with a sidebar and main content
    st.sidebar.title(f"👤 {st.session_state.username}")
    
    # Navigation menu
    st.sidebar.header("Navigation")
    
    # Sidebar navigation buttons
    if st.sidebar.button("📥 Tasks", key="nav_tasks", 
                        type="primary" if st.session_state.current_view == "tasks" else "secondary"):
        st.session_state.current_view = "tasks"
        st.rerun()
    
    if st.sidebar.button("📊 Reports", key="nav_reports", 
                        type="primary" if st.session_state.current_view == "reports" else "secondary"):
        st.session_state.current_view = "reports"
        st.rerun()
    
    # Auto-refresh settings
    st.sidebar.header("Settings")
    auto_refresh = st.sidebar.checkbox("Enable Auto-Refresh", value=True)
    refresh_interval = st.sidebar.slider("Refresh interval (seconds)", 5, 60, 15)
    
    # Auto-refresh logic
    current_count = get_message_count()
    current_time = time.time()
    
    if auto_refresh and (current_time - st.session_state.last_refresh > refresh_interval or current_count != st.session_state.last_count):
        st.session_state.last_refresh = current_time
        st.session_state.last_count = current_count
        time.sleep(0.1)  # Small delay to ensure the refresh happens
        st.rerun()
    
    # Manual refresh button
    if st.sidebar.button("🔄 Refresh Now"):
        st.session_state.last_count = current_count
        st.rerun()

    # ONGOING TASKS SECTION in sidebar
    st.sidebar.header("🔄 My Tasks")
    
    # Handle task completion if the completion form is submitted
    if st.session_state.show_completion_form:
        with st.sidebar.form(key="completion_form"):
            st.write(f"Complete task from: {st.session_state.completion_task.get('sender', 'Unknown')}")
            completion_note = st.text_area("Completion Note", height=100)
            submit_button = st.form_submit_button("Submit")
            cancel_button = st.form_submit_button("Cancel")
            
            if submit_button and completion_note:
                task_id = st.session_state.completion_task.get('id') or st.session_state.completion_task.get('message_id')
                if update_message_status(task_id, "completed", st.session_state.username, completion_note):
                    st.sidebar.success("Task marked as completed!")
                    # Reset the completion form state
                    st.session_state.show_completion_form = False
                    st.session_state.completion_task = None
                    time.sleep(1)  # Give user time to see the success message
                    st.rerun()
                else:
                    st.sidebar.error("Failed to mark task as completed")
            
            if cancel_button:
                st.session_state.show_completion_form = False
                st.session_state.completion_task = None
                st.rerun()
    
    # Display ongoing tasks in sidebar
    try:
        db = connect_db()
        # Query for messages with 'ongoing' status that are assigned to the current user
        cursor = db.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM messages WHERE status = 'ongoing' AND assigned_to = %s",
            (st.session_state.username,)
        )
        ongoing_tasks = cursor.fetchall()
        cursor.close()
        db.close()
        
        if ongoing_tasks and len(ongoing_tasks) > 0:
            for task in ongoing_tasks:
                with st.sidebar.container():
                    st.sidebar.markdown(f"**From:** {task.get('sender', 'Unknown')}")
                    st.sidebar.markdown(f"**Message:** {task.get('message', 'No message')}")
                    if 'timestamp' in task:
                        st.sidebar.markdown(f"**Time:** {task['timestamp']}")
                    
                    # Add Complete button
                    task_id = task.get('id') or task.get('message_id')
                    if st.sidebar.button("Complete Task", key=f"complete_{task_id}"):
                        st.session_state.show_completion_form = True
                        st.session_state.completion_task = task
                        st.rerun()
                    
                    st.sidebar.markdown("---")
        else:
            st.sidebar.info("No ongoing tasks")
            
    except Exception as e:
        st.sidebar.error(f"Error fetching ongoing tasks: {e}")

    # MAIN CONTENT AREA - Conditional based on current view
    if st.session_state.current_view == "tasks":
        display_tasks_view()
    elif st.session_state.current_view == "reports":
        display_reports_view()

def display_tasks_view():
    """Display the tasks view with new messages and completed tasks"""
    st.title("📨 Telegram Message Tasks")
    
    # Create tabs for New Messages and Completed Tasks
    tab1, tab2 = st.tabs(["📥 New Messages", "✅ Completed Tasks"])
    
    # NEW MESSAGES TAB
    with tab1:
        try:
            db = connect_db()
            cursor = db.cursor(dictionary=True)
            cursor.execute(
                "SELECT * FROM messages WHERE status != 'ongoing' AND status != 'completed' OR status IS NULL ORDER BY timestamp DESC"
            )
            messages = cursor.fetchall()
            cursor.close()
            db.close()
            
            if messages and len(messages) > 0:
                for msg in messages:
                    with st.container():
                        cols = st.columns([3, 1])
                        with cols[0]:
                            st.markdown(f"**From:** {msg.get('sender', 'Unknown')}")
                            st.markdown(f"**Message:** {msg.get('message', 'No message')}")
                            st.markdown(f"**Status:** {msg.get('status', 'New')}")
                            if 'timestamp' in msg:
                                st.markdown(f"**Time:** {msg['timestamp']}")
                        with cols[1]:
                            msg_id = msg.get('id') or msg.get('message_id')  # Try both possible ID field names
                            if msg_id and st.button("Accept", key=f"accept_{msg_id}"):
                                if update_message_status(msg_id, "ongoing", st.session_state.username):
                                    st.success("Task accepted!")
                                    st.rerun()
                                else:
                                    st.error("Failed to accept task")
                        st.markdown("---")
            else:
                st.info("No new messages available")
                
        except Exception as e:
            st.error(f"Error fetching messages: {e}")
    
    # COMPLETED TASKS TAB
    with tab2:
        try:
            db = connect_db()
            # Query for all completed messages for an overview
            cursor = db.cursor(dictionary=True)
            cursor.execute(
                "SELECT * FROM messages WHERE status = 'completed' ORDER BY updated_at DESC LIMIT 50"
            )
            completed_tasks = cursor.fetchall()
            cursor.close()
            db.close()
            
            # Add filters for viewing completed tasks
            filter_col1, filter_col2 = st.columns([1, 2])
            with filter_col1:
                filter_user = st.selectbox(
                    "Filter by user:",
                    ["All Users"] + list(set([task.get('assigned_to', 'Unknown') for task in completed_tasks if task.get('assigned_to')])),
                    index=0
                )
            
            with filter_col2:
                filter_date = st.date_input(
                    "Filter by date:",
                    value=datetime.now().date(),
                    key="filter_date_completed"
                )
            
            # Apply filters
            filtered_tasks = completed_tasks
            if filter_user != "All Users":
                filtered_tasks = [task for task in filtered_tasks if task.get('assigned_to') == filter_user]
            
            if filter_date:
                filtered_tasks = [
                    task for task in filtered_tasks if 
                    task.get('updated_at') and 
                    task.get('updated_at').date() == filter_date
                ]
            
            if filtered_tasks and len(filtered_tasks) > 0:
                # Show count of tasks displayed
                st.info(f"Showing {len(filtered_tasks)} completed tasks")
                
                # Display tasks
                for task in filtered_tasks:
                    with st.expander(f"{task.get('sender', 'Unknown')}: {task.get('message', 'No message')[:30]}{'...' if len(task.get('message', 'No message')) > 30 else ''}"):
                        st.markdown(f"**From:** {task.get('sender', 'Unknown')}")
                        st.markdown(f"**Message:** {task.get('message', 'No message')}")
                        st.markdown(f"**Completed by:** {task.get('assigned_to', 'Unknown')}")
                        if 'updated_at' in task:
                            st.markdown(f"**Completed at:** {task['updated_at']}")
                        if 'time_taken' in task and task['time_taken']:
                            st.markdown(f"**Time taken:** {task['time_taken']} minutes")
                        st.markdown(f"**Note:** {task.get('completion_note', 'No note provided')}")
            else:
                st.info("No completed tasks match your filters")
                
        except Exception as e:
            st.error(f"Error fetching completed tasks: {e}")

def display_reports_view():
    """Display the reports view with activity data and download options"""
    st.title("📊 Activity Reports")
    
    # Report controls
    st.subheader("Generate Report")
    
    # Report configuration options
    col_config1, col_config2 = st.columns([1, 1])
    
    with col_config1:
        report_type = st.selectbox(
            "Select Report Period",
            ["Today", "Last 7 Days", "Last 30 Days", "Custom Range"]
        )
    
    date_range = None
    if report_type == "Today":
        date_range = "today"
        st.write(f"Report for: {datetime.now().strftime('%Y-%m-%d')}")
    elif report_type == "Last 7 Days":
        date_range = "week"
        st.write(f"Report for: {(datetime.now().date() - timedelta(days=7)).strftime('%Y-%m-%d')} to {datetime.now().strftime('%Y-%m-%d')}")
    elif report_type == "Last 30 Days":
        date_range = "month"
        st.write(f"Report for: {(datetime.now().date() - timedelta(days=30)).strftime('%Y-%m-%d')} to {datetime.now().strftime('%Y-%m-%d')}")
    elif report_type == "Custom Range":
        with col_config2:
            custom_dates = st.date_input(
                "Select Date Range",
                [datetime.now().date() - timedelta(days=7), datetime.now().date()],
                format="YYYY-MM-DD"
            )
            if isinstance(custom_dates, tuple) or isinstance(custom_dates, list):
                if len(custom_dates) == 2:
                    date_range = f"{custom_dates[0]} to {custom_dates[1]}"
                    st.write(f"Report for: {custom_dates[0].strftime('%Y-%m-%d')} to {custom_dates[1].strftime('%Y-%m-%d')}")
    
    # Generate report button
    if st.button("Generate Report"):
        st.session_state.generated_report = True
        st.session_state.report_date_range = date_range
        st.rerun()
    
    # Display report if generated
    if 'generated_report' in st.session_state and st.session_state.generated_report:
        st.subheader("Activity Report")
        
        # Get activity data
        activities = get_activity_data(st.session_state.report_date_range)
        
        if activities:
            # Create DataFrame for display and download
            df = pd.DataFrame(activities)
            
            # Format columns for better display
            df_display = create_report_preview(activities)
            
            # Show summary stats
            st.subheader("Summary")
            total_tasks = len(df)
            total_time = df['time_taken'].sum() if 'time_taken' in df.columns and df['time_taken'].notna().any() else 0
            
            # Display stats in metrics
            col_stats1, col_stats2, col_stats3 = st.columns(3)
            col_stats1.metric("Total Tasks Completed", total_tasks)
            col_stats2.metric("Total Time (minutes)", int(total_time))
            
            # Get assignments by user
            if 'assigned_to' in df.columns:
                by_user = df.groupby('assigned_to').size().reset_index(name='tasks')
                by_user.columns = ['User', 'Tasks Completed']
                col_stats3.metric("Unique Users", len(by_user))
                
                # Calculate time by user if available
                if 'time_taken' in df.columns:
                    time_by_user = df.groupby('assigned_to')['time_taken'].sum().reset_index()
                    time_by_user.columns = ['User', 'Total Minutes']
                    
                    # Merge with count data
                    by_user = pd.merge(by_user, time_by_user, on='User', how='left')
                    by_user['Avg Minutes/Task'] = round(by_user['Total Minutes'] / by_user['Tasks Completed'], 1)
                
                # Show user breakdown
                st.subheader("Tasks by User")
                st.dataframe(by_user)
            
            # Show report preview
            st.subheader("Report Preview")
            st.dataframe(df_display)
            
            # Download options
            st.subheader("Download Report")
            col_dl1, col_dl2 = st.columns(2)
            
            # Excel download
            df_excel = df.copy()
            # Format datetime objects for Excel
            if 'timestamp' in df_excel.columns:
                df_excel['timestamp'] = pd.to_datetime(df_excel['timestamp'])
            if 'updated_at' in df_excel.columns:
                df_excel['updated_at'] = pd.to_datetime(df_excel['updated_at'])
            
            excel_link = get_download_link(df_excel, f"activity_report_{datetime.now().strftime('%Y%m%d')}.xlsx", 'xlsx')
            col_dl1.markdown(excel_link, unsafe_allow_html=True)
            
            # PDF download
            # Select only the most important columns for PDF
            pdf_cols = ['sender', 'message', 'assigned_to', 'updated_at', 'time_taken']
            pdf_cols = [col for col in pdf_cols if col in df.columns]
            df_pdf = df[pdf_cols].copy()
            
            # Rename columns for better readability in PDF
            column_names = {
                'sender': 'Sender',
                'message': 'Message',
                'assigned_to': 'Completed By',
                'updated_at': 'Completion Time',
                'time_taken': 'Minutes Taken'
            }
            df_pdf.rename(columns={k: v for k, v in column_names.items() if k in df_pdf.columns}, inplace=True)
            
            # Limit message length for PDF
            if 'Message' in df_pdf.columns:
                df_pdf['Message'] = df_pdf['Message'].apply(lambda x: (x[:50] + '...') if isinstance(x, str) and len(x) > 50 else x)
            
            pdf_link = get_download_link(df_pdf, f"activity_report_{datetime.now().strftime('%Y%m%d')}.pdf", 'pdf')
            col_dl2.markdown(pdf_link, unsafe_allow_html=True)
        else:
            st.warning("No activity data found for the selected period")

def main():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    
    if st.session_state.logged_in:
        dashboard()
    else:
        login()

if __name__ == "__main__":
    main()